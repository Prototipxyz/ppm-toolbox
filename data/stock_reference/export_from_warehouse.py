"""
export_from_warehouse.py

Reads PPM_Warehouse_1.xlsx (Sheets & Plates + Bars, Tubes & Profiles + Fittings)
and writes lean, per-category JSON files for the iLogic macro to consume for
dimensional cross-checking (OQ-125, OQ-129, OQ-130).

Only dimensional/grade-availability data is exported -- NOT weight formulas,
NOT pricing. The macro cross-checks geometry, it doesn't cost anything.

Usage:  python3 export_from_warehouse.py PPM_Warehouse_1.xlsx data/stock_reference/
"""
import openpyxl, json, sys, os

def load(path):
    return openpyxl.load_workbook(path, data_only=True)

def table_rows(ws, hdr_row, split_col_start, split_col_end, grade_col_start, grade_headers):
    """Read a table: label in col A, split-dim columns [split_col_start..split_col_end],
       then grade columns starting at grade_col_start (only those with a numeric value =
       'available in this grade'). Stops at first row where col A is None, OR where col A
       is a long free-text note (e.g. 'Formula: ...') rather than a real dimension label --
       detected as: label is a string longer than 60 chars, which no real dimension label is."""
    out = []
    r = hdr_row + 1
    while True:
        label = ws.cell(row=r, column=1).value
        if label is None:
            break
        if isinstance(label, str) and len(label) > 60:
            break
        rec = {"label": label}
        for c in range(split_col_start, split_col_end + 1):
            key = ws.cell(row=hdr_row, column=c).value
            key = _clean_key(key)
            rec[key] = ws.cell(row=r, column=c).value
        grades = []
        for i, code in enumerate(grade_headers):
            v = ws.cell(row=r, column=grade_col_start + i).value
            if isinstance(v, (int, float)):
                grades.append(code)
        rec["grades"] = grades
        out.append(rec)
        r += 1
    return out, r

def _clean_key(h):
    if not h: return "val"
    h = str(h).lower()
    for junk in [" (mm)", "(mm)", " (m\u00b2)", "(m\u00b2)"]:
        h = h.replace(junk, "")
    return h.strip().replace(" ", "_")

def find_title_row(ws, text, start=1, end=1000):
    for r in range(start, end):
        v = ws.cell(row=r, column=1).value
        if v and text in str(v):
            return r
    raise ValueError(f"Title containing {text!r} not found")

def export_table(ws, title_text, split_cols, grade_headers, out_dir, out_name, start=1, end=1000):
    """split_cols: (start_col, end_col) 1-indexed inclusive, columns right after label col A.
       Pass split_cols=None for single-dimension tables (grades start right at column 2)."""
    title_row = find_title_row(ws, title_text, start, end)
    hdr_row = title_row + 1
    if split_cols is None:
        grade_col_start = 2
        rows, _ = table_rows(ws, hdr_row, 2, 1, grade_col_start, grade_headers)  # empty split range
    else:
        grade_col_start = split_cols[1] + 1
        rows, _ = table_rows(ws, hdr_row, split_cols[0], split_cols[1], grade_col_start, grade_headers)
    path = os.path.join(out_dir, out_name)
    json.dump(rows, open(path, 'w'), indent=1)
    print(f"  {out_name:38s} {len(rows):4d} rows")
    return rows

def main(xlsx_path, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    wb = load(xlsx_path)
    SP = wb['Sheets & Plates']
    BTP = wb['Bars, Tubes & Profiles']
    FIT = wb['Fittings']

    print("Exporting Sheets & Plates...")
    export_table(SP, "SHEET & PLATE", (2,2), ['ST_S235','ST_S355','SS_1.4301','SS_1.4404','AL_6082',
        'AL_6060','BRS_CuZn37','PL_POM','SS_1.4016','AL_1050','AL_5754','AL_5083'],
        out_dir, "sheet_plate.json")

    # Sheet Formats has Width/Length split cols at C,D (3,4)
    export_table(SP, "SHEET FORMATS", (3,4), [], out_dir, "sheet_formats.json")

    print("Exporting Bars, Tubes & Profiles...")
    export_table(BTP, "ROUND BAR", None,
        ['ST_S235','ST_S355','SS_1.4301','SS_1.4404','AL_6082','AL_6060','BRS_CuZn37','PL_POM'],
        out_dir, "round_bar.json")
    export_table(BTP, "FLAT BAR", (2,3),
        ['ST_S235','ST_S355','SS_1.4301','SS_1.4404','AL_6082','AL_6060','BRS_CuZn37','PL_POM'],
        out_dir, "flat_bar.json")
    export_table(BTP, "SQUARE BAR", None, ['ST_S235','ST_S355','SS_1.4301','SS_1.4404'],
        out_dir, "square_bar.json")
    export_table(BTP, "HEX BAR", None, ['ST_S235','ST_S355','SS_1.4301','SS_1.4404'],
        out_dir, "hex_bar.json")
    export_table(BTP, "ROUND PIPE \u2014 STRUCTURAL", (2,4),
        ['ST_S235','ST_S355','SS_1.4301','SS_1.4404','AL_6082','AL_6060','BRS_CuZn37','PL_POM'],
        out_dir, "round_pipe_structural.json")
    export_table(BTP, "THREADED PIPE", (2,6), ['ST_S235'],
        out_dir, "threaded_pipe.json")
    export_table(BTP, "ROUND TUBE \u2014 DECORATIVE", (2,4),
        ['SS_1.4301','SS_1.4401','SS_1.4404','SS_1.4571','SS_1.4016','SS_1.4372','SS_1.4509'],
        out_dir, "round_tube_decorative.json")
    export_table(BTP, "ALUMINIUM TUBE", (2,4), ['AL_6082','AL_6060'],
        out_dir, "aluminium_tube.json")
    export_table(BTP, "ALUMINIUM SQUARE/RECT TUBE", (2,6), ['AL_6060','AL_6082'],
        out_dir, "aluminium_square_rect_tube.json")
    export_table(BTP, "SHS", (2,4), ['ST_S235','ST_S355','SS_1.4301','SS_1.4404'],
        out_dir, "shs.json")
    export_table(BTP, "RHS", (2,6), ['ST_S235','ST_S355','SS_1.4301','SS_1.4404'],
        out_dir, "rhs.json")
    export_table(BTP, "EQUAL ANGLE", (2,3),
        ['ST_S235','ST_S355','SS_1.4301','SS_1.4404','AL_6060'], out_dir, "equal_angle.json")
    export_table(BTP, "UNEQUAL ANGLE", (2,4),
        ['ST_S235','ST_S355','SS_1.4301','SS_1.4404','AL_6060'], out_dir, "unequal_angle.json")
    export_table(BTP, "ALUMINIUM T-PROFILE", (2,4), ['AL_6060'],
        out_dir, "aluminium_t_profile.json")
    export_table(BTP, "RIGID SMALL TUBE", (2,4),
        ['ST_S235','SS_1.4301','SS_1.4404','BRS_CuZn37'], out_dir, "rigid_small_tube.json")

    print("Exporting Fittings...")
    fit_rows = []
    for r in range(5, FIT.max_row+1):
        source = FIT.cell(row=r, column=1).value
        if not source: continue
        rec = dict(
            source=source, category=FIT.cell(row=r,column=2).value, standard=FIT.cell(row=r,column=3).value,
            grade=FIT.cell(row=r,column=4).value, dim_label=FIT.cell(row=r,column=5).value,
            od_mm=FIT.cell(row=r,column=6).value, od2_mm=FIT.cell(row=r,column=7).value,
            wall_mm=FIT.cell(row=r,column=8).value, radius_mm=FIT.cell(row=r,column=9).value,
            length_mm=FIT.cell(row=r,column=10).value, dn=FIT.cell(row=r,column=11).value,
            pn=FIT.cell(row=r,column=12).value, thickness_mm=FIT.cell(row=r,column=13).value,
            size_inch=FIT.cell(row=r,column=14).value, size_metric_mm=FIT.cell(row=r,column=15).value,
        )
        fit_rows.append(rec)
    json.dump(fit_rows, open(os.path.join(out_dir,"fittings.json"),'w'), indent=1)
    print(f"  {'fittings.json':38s} {len(fit_rows):4d} rows")

    print("\nDone ->", out_dir)

if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else 'PPM_Warehouse_1.xlsx'
    out = sys.argv[2] if len(sys.argv) > 2 else 'data/stock_reference'
    main(xlsx, out)
